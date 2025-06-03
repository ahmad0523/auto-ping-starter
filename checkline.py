import datetime
import smtplib
from email.mime.text import MIMEText
import subprocess
import numpy as np
from openpyxl import load_workbook
import os

# Fungsi untuk membaca IP address dan parameter dari file Excel
def read_ip_from_excel(filename="line_list.xlsx"):
    file_path = os.path.join(os.path.dirname(__file__), filename)
    wb = load_workbook(file_path)
    ws = wb.active
    data_ip = []

    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):  # Asumsi data mulai dari baris kedua
        data_ip.append(row[:3])

    return data_ip



def perform_mtr(ip, count=5):
    result = subprocess.run(['mtr', '--report', '--no-dns', '--report-cycles', str(count), ip], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    output = result.stdout.decode('utf-8')
    
    lines = output.strip().splitlines()
    avg_index = None

    # Cari indeks kolom AVG
    for line in lines:
        if "Avg" in line:
            headers = line.split()
            avg_index = headers.index("Avg")
            continue

    if avg_index is None:
        return float('inf')

    for line in lines:
        if ip in line:
            parts = line.split()
            try:
                avg_rtt = float(parts[avg_index])
                return avg_rtt
            except (ValueError, IndexError):
                continue

    return float('inf')


# Fungsi untuk mengirim email dengan body HTML
def send_email(subject, body, to_email):
    sender = 'email@pengirim.com'  # Ganti dengan email pengirim
    smtpHost = 'SMTPSERVER'  # Ganti dengan SMTP server
    smtpPort = 465  # Port SSL untuk koneksi yang aman

    # Pesan email
    msg = MIMEText(body, "html")
    msg['Subject'] = f'[AutoInfra] {subject} {str(datetime.datetime.now())}'
    msg['From'] = sender
    msg['To'] = ', '.join(to_email)

    try:
        # Menghubungkan ke server SMTP menggunakan SSL
        smtpServer = smtplib.SMTP_SSL(smtpHost, smtpPort)
        smtpServer.ehlo()  # Memastikan server siap menerima koneksi

        # Login ke server SMTP dengan kredensial yang benar
        smtpServer.login(sender, "Password-email")  # Ganti dengan password yang benar

        # Mengirim email
        smtpServer.sendmail(sender, to_email, msg.as_string())

        print(f"Email berhasil dikirim ke: {', '.join(to_email)}")

    except Exception as e:
        print(f"Error: {str(e)}")

    finally:
        smtpServer.quit()


# Fungsi utama untuk memproses IP addresses dan mengirim email
def main():
    # Baca daftar IP dari file Excel
    data_ip = read_ip_from_excel("line_list.xlsx")
    
    results = []
    
    # Proses MTR untuk setiap IP
    for ip, th, linename in data_ip:
        avg_rtt = perform_mtr(ip)
        status = "OK" if avg_rtt < th else "Alert"
        results.append((ip, avg_rtt, th, linename, status))
        print(f"{ip} - Avg RTT: {avg_rtt}ms - th: {th} - linename: {linename} - Status: {status}")
    
    # Menyusun body email dalam format HTML dengan tabel
    body = """
    <html>
    <body>
    <h2>Line Status</h2>
    <table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse; text-align: center; border: 2px solid black;">
        <tr text-align: center; border: 2px solid black;>
            <th>Line Name</th>
            <th>IP Address</th>
            <th>Threshold (ms)</th>
            <th>Average RTT (ms)</th>
            <th>Status</th>
        </tr>
    """
    
    for ip, avg_rtt, th, linename, status in results:
          # Menentukan warna latar belakang berdasarkan status
        if status == "Alert":
           row_style = 'background-color: red; color: white;'  # Warna merah dengan teks putih untuk status "Alert"
        else:
           row_style = ''  # Tidak ada perubahan warna untuk status lain

        body += f"""
        <tr style="line-height: 0.5; {row_style}">
            <td>{linename}</td>
            <td>{ip}</td>
            <td>{th}</td>
            <td>{avg_rtt:.2f}</td>
            <td>{status}</td>
        </tr>
        """
    
    body += """
    </table>
    </body>
    </html>
    """

    # Kirimkan email dengan hasil
    toAdd = ['email1@penerima.com','email2@penerima.com']
    send_email("Line_Status Daily Report", body, toAdd)
    print("Email sent successfully!")

if __name__ == "__main__":
    main()
